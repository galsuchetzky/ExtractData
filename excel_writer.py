"""Write a single extracted row to a fresh .xlsx workbook."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from schema import LIST_TYPES, Schema

LONG_TEXT_FIELDS = {"chief_complaint", "medical_history_summary"}
DEFAULT_WIDTH = 22
LONG_WIDTH = 50


def _cell_value(value: Any, ftype: str) -> Any:
    if value is None:
        return ""
    if ftype in LIST_TYPES:
        try:
            return json.dumps(value, ensure_ascii=False)
        except (TypeError, ValueError):
            return str(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def write_workbook(
    row: dict[str, Any],
    schema: Schema,
    out_path: Path,
    error: str | None = None,
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "extracted"

    headers = schema.field_names()
    if error:
        headers = headers + ["_errors"]

    bold = Font(bold=True)
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = bold
        ws.column_dimensions[cell.column_letter].width = (
            LONG_WIDTH if name in LONG_TEXT_FIELDS else DEFAULT_WIDTH
        )

    type_by_name = {f.name: f.type for f in schema.fields}
    wrap = Alignment(wrap_text=True, vertical="top")
    for col_idx, name in enumerate(headers, start=1):
        if name == "_errors":
            cell = ws.cell(row=2, column=col_idx, value=error or "")
            cell.alignment = wrap
            continue
        cell = ws.cell(
            row=2,
            column=col_idx,
            value=_cell_value(row.get(name), type_by_name[name]),
        )
        cell.alignment = wrap

    wb.save(out_path)
