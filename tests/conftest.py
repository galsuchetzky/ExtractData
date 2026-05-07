"""Pytest configuration: project on sys.path, Ollama health-check fixture,
and a session-scoped pipeline-results cache so each fixture runs through the
real pipeline only once even though several tests assert against it."""
from __future__ import annotations

import json
import logging
import os
import sys
from pathlib import Path
from typing import Any

import pytest
import requests
import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def pytest_configure(config: pytest.Config) -> None:
    config.addinivalue_line(
        "markers",
        "integration: tests that hit a local Ollama instance (slow; require qwen2.5vl:7b and gemma4:26b)",
    )


@pytest.fixture(scope="session")
def ollama_host() -> str:
    return os.environ.get("OLLAMA_HOST", "http://localhost:11434")


@pytest.fixture(scope="session")
def require_ollama(ollama_host: str) -> str:
    try:
        resp = requests.get(f"{ollama_host.rstrip('/')}/api/tags", timeout=3)
        resp.raise_for_status()
    except requests.RequestException as exc:
        pytest.skip(f"Ollama not reachable at {ollama_host}: {exc}")
    names = {m.get("name") for m in resp.json().get("models", [])}
    required = {"gemma4:latest"}  # vision is Tesseract on this branch
    missing = required - names
    if missing:
        pytest.skip(f"Missing Ollama models: {', '.join(sorted(missing))}")
    return ollama_host


# --- Session-scoped pipeline fixture -------------------------------------

FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"


def _list_fixture_paths() -> list[Path]:
    if not FIXTURES_DIR.is_dir():
        return []
    return sorted(
        p
        for p in FIXTURES_DIR.iterdir()
        if p.is_dir() and (p / "expected.yaml").is_file() and any(p.glob("page_*.png"))
    )


@pytest.fixture(scope="session")
def pipeline_outputs(
    require_ollama: str,
    tmp_path_factory: pytest.TempPathFactory,
) -> dict[str, dict[str, Any]]:
    """Run the full pipeline once per fixture; cache row + transcript + expected.

    Returns: {fixture_name: {"row": dict, "transcript": str,
                             "expected": dict, "out_xlsx": Path}}
    Tests in test_pipeline.py share this so we don't pay the LLM cost three
    times per fixture (matchers + judge-extraction + judge-transcript).
    """
    import pipeline  # local import: needs ROOT on sys.path

    logging.getLogger().setLevel(logging.INFO)
    schema_path = ROOT / "schema.yaml"
    out_root = tmp_path_factory.mktemp("pipeline")
    results: dict[str, dict[str, Any]] = {}

    for fixture in _list_fixture_paths():
        out_xlsx = out_root / f"{fixture.name}.xlsx"
        transcript_path = out_root / f"{fixture.name}.txt"
        pipeline.run(
            input_folder=fixture,
            schema_path=schema_path,
            out_xlsx=out_xlsx,
            ollama_host=require_ollama,
            save_text=transcript_path,
        )
        wb = load_workbook(out_xlsx)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        values = [c.value for c in ws[2]]
        row = dict(zip(headers, values))
        # Decode JSON-serialised list/object cells back into Python objects so
        # the judge sees real lists, not strings.
        for k, v in list(row.items()):
            if isinstance(v, str) and v and v[0] in "[{":
                try:
                    row[k] = json.loads(v)
                except json.JSONDecodeError:
                    pass
        expected = yaml.safe_load(
            (fixture / "expected.yaml").read_text(encoding="utf-8")
        )
        transcript = (
            transcript_path.read_text(encoding="utf-8")
            if transcript_path.exists()
            else ""
        )
        results[fixture.name] = {
            "row": row,
            "transcript": transcript,
            "expected": expected,
            "out_xlsx": out_xlsx,
        }
    return results
