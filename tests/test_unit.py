"""Unit tests covering schema loading, Excel writing, and JSON parsing.

These do NOT require Ollama and run in well under a second.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

import excel_writer
from extract_struct import _try_parse
from schema import LIST_TYPES, load_schema

ROOT = Path(__file__).resolve().parent.parent


@pytest.fixture(scope="module")
def schema():
    return load_schema(ROOT / "schema.yaml")


def test_schema_loads(schema):
    assert schema.version == 1
    names = schema.field_names()
    assert "patient_name" in names
    assert "exposed_to_pigeons" in names
    assert len(names) == len(set(names)), "duplicate field names"


def test_schema_render_includes_aliases_and_enum_values(schema):
    rendered = schema.render_for_prompt()
    assert "exposed_to_pigeons" in rendered
    assert "[allowed: yes, no, unknown]" in rendered
    assert any(alias in rendered for alias in ("ת.ז.", "תעודת זהות"))


def test_schema_empty_row_defaults(schema):
    row = schema.empty_row()
    for f in schema.fields:
        if f.type in LIST_TYPES:
            assert row[f.name] == []
        else:
            assert row[f.name] is None


def test_excel_writer_roundtrip_with_hebrew(schema, tmp_path: Path):
    row = {
        "patient_name": "ישראל ישראלי",
        "patient_id": "123456789",
        "date_of_birth": "01/01/1980",
        "visit_date": "07/05/2026",
        "doctor_name": "Dr. Cohen",
        "clinic": "מכבי",
        "chief_complaint": "שיעול",
        "diagnosis": ["דלקת ריאות", "HP"],
        "symptoms": ["cough", "קוצר נשימה"],
        "medical_history_summary": "no history",
        "medications": [{"name": "Prednisone", "dose": "40mg", "frequency": "daily"}],
        "procedures": ["chest x-ray"],
        "referrals": ["pulmonology"],
        "exposed_to_pigeons": "yes",
    }
    out = tmp_path / "out.xlsx"
    excel_writer.write_workbook(row, schema, out)
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    values = [c.value for c in ws[2]]
    assert headers == schema.field_names()
    by_name = dict(zip(headers, values))
    assert by_name["patient_name"] == "ישראל ישראלי"
    assert by_name["clinic"] == "מכבי"
    assert json.loads(by_name["diagnosis"]) == ["דלקת ריאות", "HP"]
    assert json.loads(by_name["medications"])[0]["name"] == "Prednisone"


def test_excel_writer_appends_errors_column(schema, tmp_path: Path):
    out = tmp_path / "out.xlsx"
    excel_writer.write_workbook(schema.empty_row(), schema, out, error="boom")
    wb = load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers[-1] == "_errors"
    assert ws.cell(row=2, column=len(headers)).value == "boom"


def test_excel_writer_blanks_for_missing_values(schema, tmp_path: Path):
    row = schema.empty_row()  # all None / []
    out = tmp_path / "blank.xlsx"
    excel_writer.write_workbook(row, schema, out)
    wb = load_workbook(out)
    ws = wb.active
    by_name = dict(zip([c.value for c in ws[1]], [c.value for c in ws[2]]))
    assert by_name["patient_name"] in (None, "")
    assert by_name["diagnosis"] == "[]"  # list defaults serialize to []


@pytest.mark.parametrize(
    "raw, expected",
    [
        ('{"a": 1}', {"a": 1}),
        ('  {"a": 1}\n', {"a": 1}),
        ('here is the answer: {"a": 1, "b": "x"} thanks', {"a": 1, "b": "x"}),
        ("not json at all", None),
        ("", None),
        ("[1,2,3]", None),  # arrays at top level: not a dict
    ],
)
def test_try_parse_handles_loose_responses(raw, expected):
    assert _try_parse(raw) == expected
